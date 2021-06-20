from openpyxl import Workbook
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook

def test1_create_sheet():
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.merge_cells('A1:D1')
    active_sheet["A1"] = "hello this is header"
    active_sheet["A2"] = "another field"
    active_sheet["A3"] = "and another field"
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook

def test2_create_sheet():
    from tempfile import NamedTemporaryFile
    from openpyxl import Workbook
    wb = Workbook()
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream

def create_sheet():
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.merge_cells('A1:D1')
    active_sheet["A1"] = "hello this is header"
    active_sheet["A2"] = "another field"
    active_sheet["A3"] = "and another field"
    return save_virtual_workbook(wb)

def main():
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.merge_cells('A1:D1')
    active_sheet["A1"] = "hello this is header"
    active_sheet["A2"] = "another field"
    active_sheet["A3"] = "and another field"
    wb.save(filename="./test2.xlsx")


if __name__ == '__main__':
    main()
